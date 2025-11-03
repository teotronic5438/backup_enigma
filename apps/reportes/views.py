# apps/reportes/views.py
from django.views.generic import TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin
from apps.ingresos.models import Remitos
from apps.ordenes.models import Ordenes
from .forms import ReporteFechasForm

# Dashboard principal
class ReporteDashboardView(LoginRequiredMixin, TemplateView):
    template_name = "reportes/dashboard.html"

# Reporte de ingresos
class ReporteIngresosView(LoginRequiredMixin, TemplateView):
    template_name = "reportes/ingresos.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        form = ReporteFechasForm(self.request.GET or None)
        context['form'] = form

        if form.is_valid():
            fecha_inicio = form.cleaned_data.get('fecha_inicio')
            fecha_fin = form.cleaned_data.get('fecha_fin')
            incluir_inactivas = form.cleaned_data.get('incluir_inactivas')

            ingresos = Remitos.objects.all()
            if fecha_inicio and fecha_fin:
                ingresos = ingresos.filter(fecha_ingreso__range=[fecha_inicio, fecha_fin])
            if not incluir_inactivas:
                ingresos = ingresos.filter(aprobado=True)

            context['ingresos'] = ingresos

        return context


import csv
from django.http import HttpResponse

# EXPORTACION COMO CSV
class ExportIngresosView(LoginRequiredMixin, TemplateView):
    def get(self, request, *args, **kwargs):
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        incluir_inactivas = request.GET.get('incluir_inactivas') == 'True'

        remitos = Remitos.objects.all()
        if fecha_inicio and fecha_fin:
            remitos = remitos.filter(fecha_ingreso__range=[fecha_inicio, fecha_fin])
        if not incluir_inactivas:
            remitos = remitos.filter(aprobado=True)

        # Crear respuesta CSV
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="reporte_ingresos.csv"'

        writer = csv.writer(response)
        writer.writerow(['Número Remito', 'Número Viaje', 'Detalle Transporte', 'Depósito', 'Producto', 'Cantidad', 'Fecha Ingreso', 'Usuario', 'Estado'])

        for remito in remitos:
            detalles = remito.remitoproducto_set.all()  # <-- CORRECCIÓN
            if detalles.exists():
                for det in detalles:
                    writer.writerow([
                        remito.numero_remito,
                        remito.numero_viaje,
                        remito.detalle_transporte,
                        remito.deposito_id,
                        f"{det.producto_id.marca} - {det.producto_id.modelo}",  # <-- CORRECCIÓN
                        det.cantidad,
                        remito.fecha_ingreso,
                        remito.usuario.nombre,
                        'Aprobado' if remito.aprobado else 'Pendiente'
                    ])
            else:
                writer.writerow([
                    remito.numero_remito,
                    remito.numero_viaje,
                    remito.detalle_transporte,
                    remito.deposito_id,
                    '',
                    '',
                    remito.fecha_ingreso,
                    remito.usuario.nombre,
                    'Aprobado' if remito.aprobado else 'Pendiente'
                ])
        return response


# apps/reportes/views.py
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.views.generic import TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin
from apps.ingresos.models import Remitos

# EXPORTACION COMO EXCEL
# apps/reportes/views.py
class ExportIngresosExcelView(LoginRequiredMixin, TemplateView):
    def get(self, request, *args, **kwargs):
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        incluir_inactivas = request.GET.get('incluir_inactivas') == 'True'

        remitos = Remitos.objects.all()
        if fecha_inicio and fecha_fin:
            remitos = remitos.filter(fecha_ingreso__range=[fecha_inicio, fecha_fin])
        if not incluir_inactivas:
            remitos = remitos.filter(aprobado=True)

        # Crear libro de Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte de Ingresos"

        # Cabecera
        headers = ['Número Remito', 'Número Viaje', 'Detalle Transporte', 'Depósito', 'Producto', 'Cantidad', 'Fecha Ingreso', 'Usuario', 'Estado']
        ws.append(headers)

        # Ajustar ancho columnas
        for i, col in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(i)].width = 20

        # Filas con datos
        for remito in remitos:
            detalles = remito.remitoproducto_set.all()
            if detalles.exists():
                for det in detalles:
                    ws.append([
                        remito.numero_remito,
                        remito.numero_viaje,
                        remito.detalle_transporte,
                        str(remito.deposito_id),  # <-- CORRECCIÓN
                        f"{det.producto_id.marca} - {det.producto_id.modelo}",
                        det.cantidad,
                        remito.fecha_ingreso.strftime('%d-%m-%Y'),
                        remito.usuario.nombre,
                        'Aprobado' if remito.aprobado else 'Pendiente'
                    ])
            else:
                ws.append([
                    remito.numero_remito,
                    remito.numero_viaje,
                    remito.detalle_transporte,
                    str(remito.deposito_id),  # <-- CORRECCIÓN
                    '',
                    '',
                    remito.fecha_ingreso.strftime('%d-%m-%Y'),
                    remito.usuario.nombre,
                    'Aprobado' if remito.aprobado else 'Pendiente'
                ])

        # Preparar respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="reporte_ingresos.xlsx"'
        wb.save(response)
        return response





# a partir de aqui trabajaremos con la vista para ordenes

# apps/reportes/views.py
from django.views.generic import TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponse
import csv
from openpyxl import Workbook
from .forms import OrdenesFiltroForm
from apps.ordenes.models import Ordenes


# === REPORTE ORDENES ===
class ReporteOrdenesView(LoginRequiredMixin, TemplateView):
    template_name = "reportes/ordenes.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        form = OrdenesFiltroForm(self.request.GET or None)
        ordenes = Ordenes.objects.none()  # inicializamos con todos

        if form.is_valid():
            modelo = form.cleaned_data.get('modelo')
            estado = form.cleaned_data.get('estado')
            tipo_accion = form.cleaned_data.get('tipo_accion')
            fecha_inicio = form.cleaned_data.get('fecha_inicio')
            fecha_fin = form.cleaned_data.get('fecha_fin')

            if modelo:
                ordenes = ordenes.filter(equipo_id__producto_id__modelo__icontains=modelo)

            if estado:
                ordenes = ordenes.filter(estado_id__nombre_estado__icontains=estado)

            if tipo_accion:
                ordenes = ordenes.filter(destino__nombre_destino__icontains=tipo_accion)


            if fecha_inicio and fecha_fin:
                ordenes = ordenes.filter(fecha_creacion__range=[fecha_inicio, fecha_fin])

        # Totales por estado (ajustar IDs según tu base)
        PENDIENTE_ID = 1
        COMPLETADO_ID = 2
        AVERIA_ID = 3
        DESTRUCCION_ID = 4

        context['total_pendientes'] = ordenes.filter(estado_id=PENDIENTE_ID).count()
        context['total_completadas'] = ordenes.filter(estado_id=COMPLETADO_ID).count()
        context['total_averia_destruccion'] = ordenes.filter(estado_id__in=[AVERIA_ID, DESTRUCCION_ID]).count()

        context['ordenes'] = ordenes
        context['form'] = form
        return context


# === EXPORT CSV ===
class ExportOrdenesCSVView(LoginRequiredMixin, TemplateView):
    def get(self, request, *args, **kwargs):
        ordenes = Ordenes.objects.all()
        form = OrdenesFiltroForm(request.GET)

        if form.is_valid():
            modelo = form.cleaned_data.get('modelo')
            estado = form.cleaned_data.get('estado')
            tipo_accion = form.cleaned_data.get('tipo_accion')
            fecha_inicio = form.cleaned_data.get('fecha_inicio')
            fecha_fin = form.cleaned_data.get('fecha_fin')

            if modelo:
                ordenes = ordenes.filter(equipo_id__producto_id__modelo__icontains=modelo)

            if estado:
                ordenes = ordenes.filter(estado_id__nombre_estado__icontains=estado)

            if tipo_accion:
                ordenes = ordenes.filter(destino__nombre_destino__icontains=tipo_accion)

            if fecha_inicio and fecha_fin:
                ordenes = ordenes.filter(fecha_creacion__range=[fecha_inicio, fecha_fin])

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="reporte_ordenes.csv"'

        writer = csv.writer(response)
        writer.writerow(['ID', 'Modelo', 'Estado', 'Tipo Acción', 'Fecha Creación', 'Fecha Modificación', 'Activo', 'Usuario'])

        for o in ordenes:
            writer.writerow([
                o.id,
                o.equipo_id.modelo if o.equipo_id else '',
                o.estado_id.nombre if o.estado_id else '',
                o.tipo_accion,
                o.fecha_creacion.strftime('%d/%m/%Y %H:%M') if o.fecha_creacion else '',
                o.fecha_modificacion.strftime('%d/%m/%Y %H:%M') if o.fecha_modificacion else '',
                'Sí' if o.activo else 'No',
                o.usuario.nombre if hasattr(o, 'usuario') else ''
            ])
        return response


# === EXPORT EXCEL ===
class ExportOrdenesExcelView(LoginRequiredMixin, TemplateView):
    def get(self, request, *args, **kwargs):
        ordenes = Ordenes.objects.all()
        form = OrdenesFiltroForm(request.GET)

        if form.is_valid():
            modelo = form.cleaned_data.get('modelo')
            estado = form.cleaned_data.get('estado')
            tipo_accion = form.cleaned_data.get('tipo_accion')
            fecha_inicio = form.cleaned_data.get('fecha_inicio')
            fecha_fin = form.cleaned_data.get('fecha_fin')

            if modelo:
                ordenes = ordenes.filter(equipo_id__producto_id__modelo__icontains=modelo)

            if estado:
                ordenes = ordenes.filter(estado_id__nombre_estado__icontains=estado)

            if tipo_accion:
                ordenes = ordenes.filter(destino__nombre_destino__icontains=tipo_accion)

            if fecha_inicio and fecha_fin:
                ordenes = ordenes.filter(fecha_creacion__range=[fecha_inicio, fecha_fin])

        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Órdenes"

        headers = ['ID', 'Modelo', 'Estado', 'Tipo Acción', 'Fecha Creación', 'Fecha Modificación', 'Activo', 'Usuario']
        ws.append(headers)

        for o in ordenes:
            ws.append([
                o.id,
                o.equipo_id.modelo if o.equipo_id else '',
                o.estado_id.nombre if o.estado_id else '',
                o.tipo_accion,
                o.fecha_creacion.strftime('%d/%m/%Y %H:%M') if o.fecha_creacion else '',
                o.fecha_modificacion.strftime('%d/%m/%Y %H:%M') if o.fecha_modificacion else '',
                'Sí' if o.activo else 'No',
                o.usuario.nombre if hasattr(o, 'usuario') else ''
            ])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="reporte_ordenes.xlsx"'
        wb.save(response)
        return response
